"""Erzeugt die Amazon "An Amazon senden"-Manifest-Upload-Datei (.xlsx)
aus den extrahierten Artikelzeilen.

Format: bündelt das offizielle Amazon-Template (assets/) und füllt das Blatt
"Create workflow – template": Default prep/labeling owner + je Zeile
Merchant SKU (Artikelnr. + Suffix) und Quantity.
"""

from __future__ import annotations

import io
import os
from dataclasses import dataclass
from typing import List, Optional

import openpyxl

ASSET_DIR = os.path.join(os.path.dirname(os.path.dirname(__file__)), "assets")
TEMPLATE_PATH = os.path.join(ASSET_DIR, "ManifestFileUpload_Template_MPL.xlsx")
WORKFLOW_SHEET = "Create workflow – template"


@dataclass
class ManifestRow:
    merchant_sku: str
    quantity: int


def build_manifest_rows(items, sku_suffix: str = "-FBA") -> List[ManifestRow]:
    """Wandelt extrahierte Items in Manifest-Zeilen um (Artikelnr. + Suffix, Menge)."""
    rows: List[ManifestRow] = []
    for it in items:
        art = (it.artikelnr or "").strip()
        if not art:
            continue
        sku = art if not sku_suffix or art.endswith(sku_suffix) else f"{art}{sku_suffix}"
        rows.append(ManifestRow(merchant_sku=sku, quantity=int(it.menge or 0)))
    return rows


def _find_label_row(ws, label: str) -> Optional[int]:
    for r in range(1, ws.max_row + 1):
        v = ws.cell(r, 1).value
        if isinstance(v, str) and v.strip().lower() == label.lower():
            return r
    return None


def _find_header_row(ws) -> int:
    for r in range(1, ws.max_row + 1):
        v = ws.cell(r, 1).value
        if isinstance(v, str) and v.strip().lower() == "merchant sku":
            return r
    raise RuntimeError("Kopfzeile 'Merchant SKU' in der Vorlage nicht gefunden.")


def build_manifest_xlsx(
    rows: List[ManifestRow],
    default_prep_owner: str = "Seller",
    default_labeling_owner: str = "Seller",
) -> bytes:
    """Baut die Manifest-Datei auf Basis des Amazon-Templates und gibt sie als Bytes zurück."""
    wb = openpyxl.load_workbook(TEMPLATE_PATH)
    ws = wb[WORKFLOW_SHEET]

    prep_row = _find_label_row(ws, "Default prep owner")
    label_row = _find_label_row(ws, "Default labeling owner")
    if prep_row:
        ws.cell(prep_row, 2, default_prep_owner)
    if label_row:
        ws.cell(label_row, 2, default_labeling_owner)

    header_row = _find_header_row(ws)
    start = header_row + 1

    # vorhandene Datenzeilen leeren (falls die Vorlage Beispielzeilen enthält)
    for r in range(start, ws.max_row + 1):
        for c in range(1, 5):
            ws.cell(r, c, None)

    for i, row in enumerate(rows):
        ws.cell(start + i, 1, row.merchant_sku)
        ws.cell(start + i, 2, row.quantity)

    out = io.BytesIO()
    wb.save(out)
    return out.getvalue()

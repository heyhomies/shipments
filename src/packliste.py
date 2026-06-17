"""Befüllt die von Amazon generierte, sendungsspezifische Packliste (.xlsx)
mit Kartoninhalten (Stückzahl je SKU je Karton) und Kartonmaßen/-gewichten
aus den extrahierten Lager-Daten.

Wichtig: Amazon nummeriert die Kartons der Verpackungsgruppe eigenständig
(P1-B1, P1-B2 …). Welche Lager-Kartonnummer welcher Spalte entspricht, weiß
nur der Anwender — deshalb wird diese Zuordnung als Mapping übergeben und ist
in der App editierbar.
"""

from __future__ import annotations

import io
import re
from dataclasses import dataclass, field
from typing import Dict, List, Optional, Tuple

import openpyxl
from openpyxl.utils import get_column_letter

SHEET_PREFIX = "Verpackungsinformationen"

_META_LABELS = {
    "weight": "kartongewicht",
    "width": "kartonbreite",
    "length": "kartonlänge",
    "height": "kartonhöhe",
    "name": "kartonname",
}


@dataclass
class Packliste:
    sheet_title: str
    total_boxes: int
    first_box_col: int
    box_cols: List[int]
    sku_rows: List[Tuple[int, str, int]]  # (zeilennr, sku, erwartete_stueckzahl)
    row_name: Optional[int]
    row_weight: Optional[int]
    row_width: Optional[int]
    row_length: Optional[int]
    row_height: Optional[int]


def normalize_sku(sku: str) -> str:
    """Vereinheitlicht eine SKU für den Abgleich (entfernt -FBA / -neu Zusätze)."""
    s = (sku or "").strip().lower()
    changed = True
    while changed:
        changed = False
        for suf in ("-fba-neu", "-fba", "-neu"):
            if s.endswith(suf):
                s = s[: -len(suf)]
                changed = True
    return s


def _find_pack_sheet(wb) -> str:
    for ws in wb.worksheets:
        if ws.title.strip().lower().startswith(SHEET_PREFIX.lower()):
            return ws.title
    raise RuntimeError(f"Kein Blatt gefunden, das mit '{SHEET_PREFIX}' beginnt.")


def _find_total_boxes(ws) -> Tuple[int, int]:
    """Liefert (total_boxes, first_box_col). Sucht die Zelle 'Gesamtzahl der Kartons:'."""
    for r in range(1, min(ws.max_row, 10) + 1):
        for c in range(1, ws.max_column + 1):
            v = ws.cell(r, c).value
            if isinstance(v, str) and "gesamtzahl der kartons" in v.lower():
                # Wert steht rechts davon
                for cc in range(c + 1, ws.max_column + 1):
                    val = ws.cell(r, cc).value
                    if isinstance(val, (int, float)):
                        return int(val), cc
    # Fallback: Standardlayout
    return 0, 13


def _find_header_row(ws) -> int:
    for r in range(1, ws.max_row + 1):
        v = ws.cell(r, 1).value
        if isinstance(v, str) and v.strip().lower() == "sku":
            return r
    raise RuntimeError("Kopfzeile 'SKU' im Packlisten-Blatt nicht gefunden.")


def _find_meta_rows(ws) -> Dict[str, Optional[int]]:
    rows: Dict[str, Optional[int]] = {k: None for k in _META_LABELS}
    for r in range(1, ws.max_row + 1):
        v = ws.cell(r, 1).value
        if not isinstance(v, str):
            continue
        low = v.strip().lower()
        for key, label in _META_LABELS.items():
            if low.startswith(label):
                rows[key] = r
    return rows


def parse_packliste(raw_bytes: bytes) -> Packliste:
    """Liest Struktur (SKUs, Kartonspalten, Metazeilen) aus der hochgeladenen Packliste."""
    wb = openpyxl.load_workbook(io.BytesIO(raw_bytes))  # Formeln bleiben erhalten
    title = _find_pack_sheet(wb)
    ws = wb[title]

    total_boxes, first_box_col = _find_total_boxes(ws)
    header_row = _find_header_row(ws)
    meta = _find_meta_rows(ws)

    if not total_boxes:
        # aus Metazeile ableiten, falls Zähler fehlt
        total_boxes = 0

    # SKU-Zeilen einsammeln (ab header+1, bis Leerzeile / Metabereich)
    sku_rows: List[Tuple[int, str, int]] = []
    meta_min = min([r for r in meta.values() if r], default=ws.max_row + 1)
    expected_col = _find_expected_col(ws, header_row)
    for r in range(header_row + 1, meta_min):
        v = ws.cell(r, 1).value
        if v is None or not str(v).strip():
            continue
        sku = str(v).strip()
        exp = ws.cell(r, expected_col).value if expected_col else None
        exp_int = int(exp) if isinstance(exp, (int, float)) else 0
        sku_rows.append((r, sku, exp_int))

    box_cols = [first_box_col + i for i in range(total_boxes)] if total_boxes else []

    return Packliste(
        sheet_title=title,
        total_boxes=total_boxes,
        first_box_col=first_box_col,
        box_cols=box_cols,
        sku_rows=sku_rows,
        row_name=meta["name"],
        row_weight=meta["weight"],
        row_width=meta["width"],
        row_length=meta["length"],
        row_height=meta["height"],
    )


def _find_expected_col(ws, header_row: int) -> Optional[int]:
    for c in range(1, ws.max_column + 1):
        v = ws.cell(header_row, c).value
        if isinstance(v, str) and v.strip().lower().startswith("erwartete"):
            return c
    return None


def units_by_sku_box(shipment) -> Dict[str, Dict[int, int]]:
    """Aggregiert je normalisierter SKU die Stückzahl pro Lager-Kartonnummer."""
    result: Dict[str, Dict[int, int]] = {}
    for it in shipment.items:
        base = normalize_sku(it.artikelnr)
        bucket = result.setdefault(base, {})
        for k in it.kartons:
            bucket[k.nummer] = bucket.get(k.nummer, 0) + int(k.stueck)
    return result


def boxes_by_number(shipment) -> Dict[int, "object"]:
    return {b.nummer: b for b in shipment.boxes}


@dataclass
class FillSuggestion:
    # units[(zeilennr, box_number)] = stueck
    units: Dict[Tuple[int, int], int] = field(default_factory=dict)
    # boxmeta[box_number] = (weight, length, width, height)
    boxmeta: Dict[int, Tuple[float, float, float, float]] = field(default_factory=dict)
    unmatched_skus: List[str] = field(default_factory=list)
    missing_box_meta: List[int] = field(default_factory=list)


def suggest_fill(pack: Packliste, shipment, col_to_boxnum: Dict[int, int]) -> FillSuggestion:
    """Erzeugt einen Vorbefüllungs-Vorschlag anhand der Spalten->Kartonnummer-Zuordnung."""
    sku_units = units_by_sku_box(shipment)
    box_meta = boxes_by_number(shipment)
    box_numbers = set(col_to_boxnum.values())

    sug = FillSuggestion()
    for (row, sku, _exp) in pack.sku_rows:
        base = normalize_sku(sku)
        bucket = sku_units.get(base)
        if bucket is None:
            sug.unmatched_skus.append(sku)
            continue
        for boxnum in box_numbers:
            qty = bucket.get(boxnum, 0)
            if qty:
                sug.units[(row, boxnum)] = qty

    for boxnum in sorted(box_numbers):
        b = box_meta.get(boxnum)
        if b is None:
            sug.missing_box_meta.append(boxnum)
        else:
            sug.boxmeta[boxnum] = (b.gewicht_kg, b.laenge_cm, b.breite_cm, b.hoehe_cm)
    return sug


def write_packliste(
    raw_bytes: bytes,
    pack: Packliste,
    col_to_boxnum: Dict[int, int],
    units: Dict[Tuple[int, int], int],
    boxmeta: Dict[int, Tuple[float, float, float, float]],
) -> bytes:
    """Schreibt Stückzahlen und Kartonmaße/-gewichte in die Packliste und gibt Bytes zurück.

    - units: {(zeilennr, box_number): stueck}
    - boxmeta: {box_number: (gewicht_kg, laenge_cm, breite_cm, hoehe_cm)}
    Amazon-Formeln (z.B. 'Stückzahl im Karton', Kartonnamen) bleiben unberührt.
    """
    wb = openpyxl.load_workbook(io.BytesIO(raw_bytes))
    ws = wb[pack.sheet_title]

    for col, boxnum in col_to_boxnum.items():
        # Stückzahlen je SKU-Zeile
        for (row, sku, _exp) in pack.sku_rows:
            qty = units.get((row, boxnum), 0)
            ws.cell(row, col, qty if qty else None)
        # Maße / Gewicht
        meta = boxmeta.get(boxnum)
        if meta:
            weight, length, width, height = meta
            if pack.row_weight:
                ws.cell(pack.row_weight, col, weight)
            if pack.row_width:
                ws.cell(pack.row_width, col, width)
            if pack.row_length:
                ws.cell(pack.row_length, col, length)
            if pack.row_height:
                ws.cell(pack.row_height, col, height)

    out = io.BytesIO()
    wb.save(out)
    return out.getvalue()
